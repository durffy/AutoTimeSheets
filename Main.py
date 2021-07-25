import os

import openpyxl
from openpyxl import Workbook
from pathlib import Path

from Util.TimeSheet import FortyHourTimeSheet

#for each workbook in a month
    #create a new workbook
    #for each pay period
        # create a new worksheet
        #for each employee
            #get the time worked
            #calculate time left over from 8 hour shifts
            #auto generate time sheets in a single sheet




def parseNames(destination_ws, names, hours, date):
    weekdays = ["Mon", "Tues", "Wed", "Thurs", "Fri", "Sat"]
    for i in range(len(names)):
            FortyHourTimeSheet(destination_ws, names[i], date, hours[i])


def parseFiles(month):
    files = os.listdir("files/")
    for file_name in files:

        path = "files/" + file_name
        destination_file = str(month) + ".xlsx"

        if (os.path.exists(destination_file)):
            destination_wb = openpyxl.load_workbook(destination_file)
        else:
            destination_wb = Workbook()

        if (str(month) == file_name.split("-")[0]):

            if (len(file_name) >= len("9-06-19 (1).xlsx")):
                date = file_name[:len(file_name) - 9]
            else:
                date = file_name[:len(file_name) - 5]
            destination_ws = destination_wb.create_sheet(date)

            source_file = Path(path)
            source_wb = openpyxl.load_workbook(source_file)
            source_ws = source_wb["Sheet1"]

            # all the values in row 1
            names = []
            hours = []
            for row in source_ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if (cell.value != None):
                        names.append(cell.value)

            # all the values in row 5
            i = 5
            while (len(hours) == 0):

                for row in source_ws.iter_rows(min_row=source_ws.max_row, max_row=source_ws.max_row):
                    for cell in row:
                        if (cell.value != None):
                            hours.append(cell.value)

            parseNames(destination_ws, names, hours, date)
            destination_wb.save(filename=destination_file)

def main():
    months = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    for month in months:
        parseFiles(month)

if __name__=="__main__":
    main()