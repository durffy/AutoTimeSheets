import os

import openpyxl
from openpyxl import Workbook
from pathlib import Path

from ExtraTimeEntry import ExtraTime
from TimeEntry import TimeEntry

#for each workbook in a month
    #create a new workbook
    #for each pay period
        # create a new worksheet
        #for each employee
            #get the time worked
            #calculate time left over from 8 hour shifts
            #auto generate time sheets in a single sheet


def setTimeSheetHeader(destination_ws, name, date):
    destination_ws.append(["Name", name])
    destination_ws.append(["Date", date])
    destination_ws.append(["Day", "In", "Out", "Lunch", "In", "Out", "Total"])

def setDefaultTimes():
    defaultTimeEntry = TimeEntry()
    defaultTimeEntry.setDefaults()
    return defaultTimeEntry


def setTimeEntry(weekday,destination_ws, name, date, total):
    defaultTimes = setDefaultTimes()
    defaultTimeList = []

    defaultTimeList.append(weekday)
    for e in defaultTimes.getTimeList():
        defaultTimeList.append(e)
    defaultTimeList.append(total)
    destination_ws.append(defaultTimeList)
    destination_ws.append([""])


def parseNames(destination_ws, names, hours, date):
    weekdays = ["Mon", "Tues", "Wed", "Thurs", "Fri", "Sat"]
    for i in range(len(names)):

        if (names[i].lower() != "TOTAL".lower()):
            setTimeSheetHeader(destination_ws, names[i], date)
            try:
                extra_time = float(hours[i] % 8)
                workday = int((hours[i] - extra_time) / 8)
            except:
                print("issue with time")

            for day in range(workday):
                if (workday > 6):
                    print("!!! CHECK HOURS FOR:", names[i], date, hours[i])
                else:
                    setTimeEntry(weekdays[day], destination_ws, names[i], date, total=8)

            CalculatedExtraTime = ExtraTime(extra_time)
            extraTimeEntry = CalculatedExtraTime.getTimeEntry()
            extraTimeList = []

            if (extra_time != 0):
                if (day < 5):
                    day += 1
                if (workday >= 6):
                    print("!!! CHECK HOURS FOR: ")
                else:
                    extraTimeList.append(weekdays[day])
                    for e in extraTimeEntry.getTimeList():
                        extraTimeList.append(e)
                    extraTimeList.append(extra_time)
                    destination_ws.append(extraTimeList)
                    destination_ws.append([""])

            # fill out remaining days left in the workweek as zero days on timesheet
            for remaining_day in range(day + 1, len(weekdays)):
                if (workday > 6):
                    print("!!! CHECK HOURS FOR:", names[i], date, hours[i])
                else:
                    setTimeEntry(weekdays[remaining_day], destination_ws, names[i], date, total="")

            try:
                destination_ws.append(["", "", "", "", "", "Total", hours[i]])
                destination_ws.append([""])
            except:
                print("Hour len:", len(hours), "Iteration:", i)
                print("Names len:", len(names))
                print("names", names)

        elif (names[i].lower() == "TOTAL".lower()):
            total_hours = 0
            for i in range(len(hours) - 1):
                if(isinstance(hours[i], float) or isinstance(hours[i],int)):
                    total_hours += float(hours[i])
            destination_ws.append(["", "", "", "", "", "Total", total_hours])




def parseFiles(month):
    files = os.listdir("files/")
    for file_name in files:

        path = "files/" + file_name
        destination_file = str(month) + ".xlsx"

        if (os.path.exists(destination_file)):
            destination_wb = openpyxl.load_workbook(destination_file)
        else:
            destination_wb = Workbook()

        if (str(month) == file_name.split("-")[0]):

            if (len(file_name) >= len("9-06-19 (1).xlsx")):
                date = file_name[:len(file_name) - 9]
            else:
                date = file_name[:len(file_name) - 5]
            destination_ws = destination_wb.create_sheet(date)

            source_file = Path(path)
            source_wb = openpyxl.load_workbook(source_file)
            source_ws = source_wb["Sheet1"]

            # all the values in row 1
            names = []
            hours = []
            for row in source_ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if (cell.value != None):
                        names.append(cell.value)

            # all the values in row 5
            i = 5
            while (len(hours) == 0):

                for row in source_ws.iter_rows(min_row=source_ws.max_row, max_row=source_ws.max_row):
                    for cell in row:
                        if (cell.value != None):
                            hours.append(cell.value)

            parseNames(destination_ws, names, hours, date)
            destination_wb.save(filename=destination_file)

def main():
    months = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    for month in months:
        parseFiles(month)

if __name__=="__main__":
    main()